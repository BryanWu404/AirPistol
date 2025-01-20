import sys
import time
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtGui import QMouseEvent
from PyQt5.QtCore import QPointF
import pyqtgraph as pg
import numpy as np
from collections import deque
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QPushButton
from pyqtgraph import LegendItem, PlotDataItem
from PyQt5.QtWidgets import QGraphicsLineItem
import zipfile


class PressureSensorApp(QMainWindow):
    def __init__(self, *args, **kwargs):
        super(PressureSensorApp, self).__init__(*args, **kwargs)
        self.graphWidget = pg.PlotWidget()
        self.setCentralWidget(self.graphWidget)

        # 在初始化部分添加一個成員變量來存儲垂直線的參考
        self.vertical_line = None

        # 設定圖表標題、軸標籤
        title_font = QFont()
        title_font.setPointSize(40)
        title_font.setBold(True)
        self.graphWidget.setTitle("即時空氣鎗擊發壓力數據")
        axis_font = QFont()
        axis_font.setPointSize(16)
        axis_font.setBold(True)
        self.graphWidget.setLabel('left', 'Pressure')
        self.graphWidget.setLabel('bottom', 'Time', 's')

        self.time = deque(maxlen=15)  # 保持最近3秒的時間軸，每0.2秒一個數據點，共25個數據點
        self.握力 = deque(maxlen=15)  # 保持最近3秒的壓力數據，每0.2秒一個數據點，共25個數據點
        self.擊發力 = deque(maxlen=15)  # 第二個感測器的壓力數據
        self.data_line1 = self.graphWidget.plot(pen='r')
        self.data_line2 = self.graphWidget.plot(pen='g')

        # 添加圖例
        self.legend = LegendItem(offset=(70, 30))  # 圖例的位置
        self.legend.setParentItem(self.graphWidget.graphicsItem())  # 將圖例添加到圖表中
        self.legend.addItem(self.data_line1, '擊發力')
        self.legend.addItem(self.data_line2, '握力')

        # 表格用來顯示前3秒的數據
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(3)
        self.tableWidget.setColumnCount(15)
        self.tableWidget.setVerticalHeaderLabels(["Time", "握力", "擊發力"])

        # 將圖表和表格放在一個垂直佈局中
        layout = QVBoxLayout()
        layout.addWidget(self.graphWidget)
        layout.addWidget(self.tableWidget)

        # 添加按鈕
        self.exportButton = QPushButton("Export Excel")
        self.exportButton.clicked.connect(self.export_excel)
        layout.addWidget(self.exportButton)

        # 添加stretch使表格在窗口中垂直置中
        layout.addStretch(1)

        # 將垂直佈局設置為中央窗口的佈局
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        # 更新數據的定時器
        self.timer = QTimer()
        self.timer.setInterval(200)  # 每0.2秒更新一次
        self.timer.timeout.connect(self.update_plot_data)
        self.timer.start()

        # 記錄程式啟動的時間
        self.start_time = time.time()

        self.excel_filename = None  # 初始化Excel檔案名稱
        self.click_count = 0  # 初始化點擊次數

    def mousePressEvent(self, event):

        if event.button() == Qt.LeftButton:
            self.click_count += 1
            if self.excel_filename is None:
                # self.excel_filename = datetime.now().strftime("%Y%m%d%H%M") + '.xlsx'
                self.excel_filename = datetime.now().strftime("%m%d%H%M") + '.xlsx'
                self.save_data_to_excel()
            else:
                self.save_data_to_excel()
        # 繪製垂直線
        x_data = list(self.time)[-1]
        line = pg.InfiniteLine(pos=x_data, angle=90, pen='b')
        self.graphWidget.addItem(line)
        # 檢查是否有垂直線在10秒之前的位置，如果有，則移除
        for item in self.graphWidget.items():
            if isinstance(item, pg.InfiniteLine):
                if abs(item.value() - x_data) > 10:
                    self.graphWidget.removeItem(item)

    def save_data_to_excel(self):
        data = {self.click_count: list(self.握力),
                "               ": list(self.擊發力)}
        df = pd.DataFrame(data)
    # 轉置DataFrame以適應水平布局
        df = df.transpose()
        # 只保留最近3秒的資料
        df = df.iloc[-5:]

    # 如果檔案不存在，則創建新檔案
        if not os.path.exists(self.excel_filename):
            with pd.ExcelWriter(self.excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1',
                            startrow=0, index=True)
        else:
            # 打開現有檔案
            book = load_workbook(self.excel_filename)
            writer = pd.ExcelWriter(self.excel_filename, engine='openpyxl')
            # writer.book = book
            writer.workbook = book
            # writer.sheets = {ws.title: ws for ws in book.worksheets}
            for ws in book.worksheets:
                writer.book.create_sheet(title=ws.title)
                writer.book[ws.title] = ws

        # 獲取要開始寫入的新列的索引
            if 'Sheet1' in book.sheetnames:
                sheet = book['Sheet1']
                max_row = sheet.max_row
            # 將DataFrame轉置並寫入新的列
                df.to_excel(writer, sheet_name='Sheet1',
                            startrow=max_row + 1, index=True, header=False)
            writer.close()

    def update_plot_data(self):
        current_time = time.time()
        elapsed_time = current_time - self.start_time

        self.time.append(elapsed_time)
        self.握力.append(np.random.normal(10, 2))  # 使用隨機數據模擬第一個感測器的壓力讀數
        self.擊發力.append(np.random.uniform(20, 30))  # 使用隨機數據模擬第二個感測器的壓力讀數

        # 保持time列表中只有最近10秒的數據
        while self.time and self.time[-1] - self.time[0] > 10:
            self.time.pop(0)
            self.握力.pop(0)
            self.擊發力.pop(0)

        # 更新圖表數據
        self.data_line1.setData(list(self.time), list(self.握力))
        self.data_line2.setData(list(self.time), list(self.擊發力))

        # 更新圖表的x軸範圍以顯示最近10秒的數據
        self.graphWidget.setXRange(min(self.time), max(self.time), padding=0)

        # 更新表格的數據，只顯示最近3秒的數據
        for i in range(min(len(self.time), 15)):
            self.tableWidget.setItem(0, i, QTableWidgetItem(
                str(round(self.time[-i-1], 2))))
            self.tableWidget.setItem(
                1, i, QTableWidgetItem(str(round(self.握力[-i-1], 2))))
            self.tableWidget.setItem(
                2, i, QTableWidgetItem(str(round(self.擊發力[-i-1], 2))))

    def export_excel(self):
        if self.excel_filename:
            os.startfile(os.path.abspath(self.excel_filename))
        else:
            print("Excel file not found.")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = PressureSensorApp()
    main.show()
    sys.exit(app.exec_())

